from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Dict, List, Any
from difflib import get_close_matches

from openpyxl import load_workbook

from .core import ComponentFlow, ElementalComposition, parse_empirical_formula
from .excel_utils import to_bool, to_float, get_header_map
from .input_utils import is_free_water_entry


@dataclass(frozen=True)
class ExcelInputData:
    global_parameters: Dict[str, Any]
    components: List[ComponentFlow]
    database: Dict[str, ElementalComposition]
    water_available_mol: float = 0.0

def read_component_database(wb) -> Dict[str, ElementalComposition]:
    if "03_COMPONENT_DATABASE" not in wb.sheetnames:
        return {}

    ws = wb["03_COMPONENT_DATABASE"]
    header_row = None
    for row in range(1, ws.max_row + 1):
        row_values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        normalized = [str(v).strip() if v is not None else "" for v in row_values]
        if "component_name" in normalized and "formula" in normalized:
            header_row = row
            break

    if header_row is None:
        return {}

    headers = get_header_map(ws, header_row)
    database: Dict[str, ElementalComposition] = {}

    for row in range(header_row + 1, ws.max_row + 1):
        name = ws.cell(row=row, column=headers["component_name"]).value
        if name is None or str(name).strip() == "":
            continue

        active_col = headers.get("active")
        active = to_bool(ws.cell(row=row, column=active_col).value, default=True) if active_col else True
        if not active:
            continue

        formula = ws.cell(row=row, column=headers["formula"]).value
        if formula is None or str(formula).strip() == "":
            continue

        database[str(name).strip().upper()] = parse_empirical_formula(str(formula).strip())

    return database


def read_biopotgas_excel(path: str | Path, sheet_name: str = "01_INPUTS") -> ExcelInputData:
    path = Path(path)
    wb = load_workbook(path, data_only=True)

    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet {sheet_name!r} was not found in {path.name}.")

    database = read_component_database(wb)
    ws = wb[sheet_name]

    global_parameters: Dict[str, Any] = {}
    for row in range(4, 18):
        key = ws[f"A{row}"].value
        value = ws[f"B{row}"].value
        if key:
            global_parameters[str(key).strip()] = value

    header_row = None
    headers: Dict[str, int] = {}
    for row in range(1, ws.max_row + 1):
        row_values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        normalized = [str(v).strip() if v is not None else "" for v in row_values]
        if "component_name" in normalized and (
            "input_quantity" in normalized or "molar_flow" in normalized
        ):
            header_row = row
            headers = {name: idx + 1 for idx, name in enumerate(normalized) if name}
            break

    if header_row is None:
        raise ValueError("The component input table header was not found.")

    required_columns = {"component_name", "degradable", "input_mode", "active_row"}
    missing = required_columns.difference(headers)
    if missing:
        raise ValueError(f"Missing required columns in input table: {sorted(missing)}")

    quantity_column = headers.get("input_quantity", headers.get("molar_flow"))
    if quantity_column is None:
        raise ValueError(
            "Missing required quantity column in input table: use 'input_quantity'."
        )

    components: List[ComponentFlow] = []
    empty_rows = 0
    water_available_mol = 0.0

    for row in range(header_row + 1, ws.max_row + 1):
        name_value = ws.cell(row=row, column=headers["component_name"]).value
        name_text = "" if name_value is None else str(name_value).strip()

        if name_text.lower().startswith("regras de preenchimento"):
            break

        row_values = [ws.cell(row=row, column=col).value for col in range(1, ws.max_column + 1)]
        if all(value is None or str(value).strip() == "" for value in row_values):
            empty_rows += 1
            if empty_rows >= 3:
                break
            continue

        empty_rows = 0

        if name_text == "":
            continue

        active = to_bool(ws.cell(row=row, column=headers["active_row"]).value, default=True)
        if not active:
            continue

        name = name_text

        quantity_value = ws.cell(row=row, column=quantity_column).value
        if quantity_value is None or str(quantity_value).strip() == "":
            raise ValueError(f"Input quantity is required for component {name!r} at row {row}.")
        
        molar_flow = to_float(quantity_value, "input_quantity", row)
        
        degradable = to_bool(ws.cell(row=row, column=headers["degradable"]).value, default=True)
        input_mode = str(ws.cell(row=row, column=headers["input_mode"]).value or "DATABASE").strip().upper()

        formula = ws.cell(row=row, column=headers["formula"]).value if "formula" in headers else None

        if is_free_water_entry(name, formula, input_mode):
            water_available_mol += molar_flow
            continue

        if input_mode == "DATABASE":
            key = name.strip().upper()
            if key not in database:
                suggestions = get_close_matches(key, database.keys(), n=3, cutoff=0.6)
                suggestion_text = (
                    f" Did you mean: {', '.join(suggestions)}?"
                    if suggestions
                    else " No similar component names were found."
                )

                raise KeyError(
                    f"Component {name!r} was not found in 03_COMPONENT_DATABASE."
                    f"{suggestion_text}"
                )
            
            composition = database[key]
            
        elif input_mode == "FORMULA":
            if formula is None or str(formula).strip() == "":
                raise ValueError(f"Formula is required for component {name!r} at row {row}.")
            composition = parse_empirical_formula(str(formula).strip())
        elif input_mode == "ELEMENTS":
            element_columns = {"C_atoms", "H_atoms", "O_atoms", "N_atoms", "S_atoms"}
            missing_elements_columns = element_columns.difference(headers)

            if missing_elements_columns:
                raise ValueError(
                    f"Missing required columns for ELEMENTS input mode: "
                    f"{sorted(missing_elements_columns)}"
                )
            
            composition = ElementalComposition(
                C=to_float(ws.cell(row=row, column=headers["C_atoms"]).value, "C_atoms", row),
                H=to_float(ws.cell(row=row, column=headers["H_atoms"]).value, "H_atoms", row),
                O=to_float(ws.cell(row=row, column=headers["O_atoms"]).value, "O_atoms", row),
                N=to_float(ws.cell(row=row, column=headers["N_atoms"]).value, "N_atoms", row),
                S=to_float(ws.cell(row=row, column=headers["S_atoms"]).value, "S_atoms", row),
            )
        else:
            raise ValueError(f"Invalid input_mode for component {name!r} at row {row}: {input_mode!r}.")

        components.append(ComponentFlow(name=name, molar_flow=molar_flow, composition=composition, degradable=degradable))

    return ExcelInputData(
        global_parameters=global_parameters,
        components=components,
        database=database,
        water_available_mol=water_available_mol,
        )
