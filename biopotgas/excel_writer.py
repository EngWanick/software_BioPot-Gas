from __future__ import annotations

from collections.abc import Mapping
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .pipeline import calculate_from_excel
from .excel_utils import get_header_map, to_bool
from .validation import compare_experimental_to_theoretical


OUTPUT_TEMPLATE_NON_RESULT_KEYS = {
    "BioPot-Gas v0.5 | Relatório Automático de Resultados",
    "BioPot-Gas v0.6 | Relatório Automático de Resultados",
    "Identificação",
    "Parâmetros usados",
    "Resultados molares",
    "Resultados mássicos",
    "Resultados volumétricos normalizados",
    "Resultados energéticos",
    "Balanço de água",
    "Mensagens e alertas",
    "Mensagens, notas e alertas",
    "Campo",
    "Project_name",
    "Analyst",
    "Date",
    "Basis_molar_unit",
}

def output_unit_for_key(key: str, results: Mapping[str, Any]) -> str | None:
    molar_keys = {
        "CH4_mol",
        "CO2_mol",
        "NH3_mol",
        "H2S_mol",
        "H2O_mol",
        "total_gas_mol",
        "inert_carbon_mol",
        "water_available_mol",
        "net_water_balance_mol",
    }

    mass_keys = {
        "CH4_mass",
        "CO2_mass",
        "NH3_mass",
        "H2S_mass",
        "H2O_mass",
        "inert_carbon_mass",
        "net_water_balance_mass",
    }

    volume_keys = {
        "CH4_Nm3",
        "CO2_Nm3",
        "NH3_Nm3",
        "H2S_Nm3",
        "total_biogas_Nm3",
        "biogas_Nm3",
    }

    energy_mj_keys = {
        "energy_MJ",
        "CH4_energy_LHV_MJ",
    }

    energy_kwh_keys = {
        "energy_kWh",
        "CH4_energy_LHV_kWh",        
    }

    percentage_keys = {
        "CH4_percent",
        "CO2_percent",
        "NH3_percent",
        "H2S_percent",
    }

    fraction_keys = {
        "CH4_fraction",
        "CO2_fraction",
        "NH3_fraction",
        "H2S_fraction",
    }

    if key in molar_keys:
        return results.get("output_molar_unit")
    if key in mass_keys:
        return results.get("output_mass_unit")
    if key in volume_keys:
        return results.get("output_volume_unit")
    if key in energy_mj_keys:
        return results.get("output_energy_MJ_unit")
    if key in energy_kwh_keys:
        return results.get("output_energy_kWh_unit")
    if key in percentage_keys:
        return "%"
    if key in fraction_keys:
        return "fraction"

    return None

def write_outputs_to_excel(
        input_path: str | Path,
        output_path: str | Path | None = None,
        results: Mapping[str, Any] | None = None,
    ) -> Path:
    input_path = Path(input_path)
    if output_path is None:
        output_path = input_path.with_name(input_path.stem + "_calculated.xlsx")
    output_path = Path(output_path)

    if results is None:
        results = calculate_from_excel(input_path)

    wb = load_workbook(input_path)
    if "02_OUTPUTS" not in wb.sheetnames:
        wb.create_sheet("02_OUTPUTS")
    ws = wb["02_OUTPUTS"]

    output_rows: dict[str, int] = {}

    for row in range(1, ws.max_row + 1):
        key = ws.cell(row=row, column=1).value
        if key is None:
            continue

        key = str(key).strip()
        if key == "":
            continue

        output_rows[key] = row

    output_template_keys = set(output_rows)

    unknown_output_keys = sorted(
        key
        for key in output_template_keys - set(results)
        if key not in OUTPUT_TEMPLATE_NON_RESULT_KEYS
    )
    unwritten_result_keys = sorted(
        key
        for key in set(results) - output_template_keys
        if key != "warnings"
    )

    output_warnings: list[str] = []
    if results.get("warnings"):
        output_warnings.append(str(results["warnings"]))

    if unwritten_result_keys:
        output_warnings.append(
            "Calculated result keys not written to 02_OUTPUTS: "
            + ", ".join(unwritten_result_keys)
            + "."
        )

    if unknown_output_keys:
        output_warnings.append(
            "02_OUTPUTS keys not found in calculated results: "
            + ", ".join(unknown_output_keys)
            + "."
        )

    if output_warnings:
        results = dict(results)
        results["warnings"] = "; ".join(output_warnings)

    for key, row in output_rows.items():
        if key in results:
            ws.cell(row=row, column=2).value = results[key]

            unit = output_unit_for_key(key, results)
            if unit is not None:
                ws.cell(row=row, column=3).value = unit

    if "04_EXPERIMENTAL_VALIDATION" in wb.sheetnames:
        val = wb["04_EXPERIMENTAL_VALIDATION"]

        header_row = None
        for row in range(1, val.max_row + 1):
            values = [val.cell(row=row, column=col).value for col in range(1, val.max_column + 1)]
            normalized = [str(v).strip() if v is not None else "" for v in values]
            if "sample_id" in normalized and "experimental_CH4_Nm3" in normalized:
                header_row = row
                break

        calc_header_row = None
        for row in range(1, val.max_row + 1):
            values = [val.cell(row=row, column=col).value for col in range(1, val.max_column + 1)]
            normalized = [str(v).strip() if v is not None else "" for v in values]
            if "absolute_error_CH4_Nm3" in normalized and "validation_status" in normalized:
                calc_header_row = row
                break

        if header_row and calc_header_row:
            h = get_header_map(val, header_row)
            ch = get_header_map(val, calc_header_row)
            output_row = calc_header_row + 1

            last_row = max(val.max_row, output_row)
            last_col = max(val.max_column, 10)

            for r in range(output_row, last_row + 1):
                for c in range(1, last_col + 1):
                    val.cell(row=r, column=c).value = None

            for r in range(header_row + 1, calc_header_row - 1):
                sample_id = val.cell(row=r, column=h["sample_id"]).value
                if sample_id is None or str(sample_id).strip() == "":
                    continue

                active_col = h.get("active_row")
                active = to_bool(val.cell(row=r, column=active_col).value, default=True) if active_col else True
                if not active:
                    continue

                theoretical_ch4 = val.cell(row=r, column=h["theoretical_CH4_Nm3"]).value
                theoretical_biogas = val.cell(row=r, column=h["theoretical_biogas_Nm3"]).value

                if theoretical_ch4 in (None, ""):
                    theoretical_ch4 = results["CH4_Nm3"]
                    val.cell(row=r, column=h["theoretical_CH4_Nm3"]).value = theoretical_ch4

                if theoretical_biogas in (None, ""):
                    theoretical_biogas = results["total_biogas_Nm3"]
                    val.cell(row=r, column=h["theoretical_biogas_Nm3"]).value = theoretical_biogas

                comp = compare_experimental_to_theoretical(
                    experimental_ch4_nm3=val.cell(row=r, column=h["experimental_CH4_Nm3"]).value,
                    experimental_biogas_nm3=val.cell(row=r, column=h["experimental_biogas_Nm3"]).value,
                    experimental_ch4_percent=val.cell(row=r, column=h["experimental_CH4_percent"]).value,
                    theoretical_ch4_nm3=float(theoretical_ch4),
                    theoretical_biogas_nm3=float(theoretical_biogas),
                )

                val.cell(row=output_row, column=1).value = sample_id
                for field, col in ch.items():
                    if field == "sample_id":
                        continue
                    if field in comp:
                        val.cell(row=output_row, column=col).value = comp[field]
                output_row += 1

    wb.save(output_path)
    return output_path