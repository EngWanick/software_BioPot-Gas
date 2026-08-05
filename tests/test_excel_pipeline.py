from pathlib import Path
import shutil

import pytest
from openpyxl import load_workbook

from biopotgas.excel_reader import (
    calculate_from_excel,
    read_biopotgas_excel,
    write_outputs_to_excel
)


TEMPLATE_FILE = Path("BioPot-Gas_Input_Template_v6_validation_benchmarks.xlsx")


def test_excel_pipeline_generates_calculated_workbook_without_overwriting_input(tmp_path):
    input_copy = tmp_path / TEMPLATE_FILE.name
    output_file = tmp_path / "calculated_output.xlsx"

    shutil.copy2(TEMPLATE_FILE, input_copy)
    original_input_bytes = input_copy.read_bytes()

    results = calculate_from_excel(input_copy)
    written_file = write_outputs_to_excel(
        input_copy,
        output_path=output_file,
        results=results,
    )

    assert written_file == output_file
    assert output_file.exists()
    assert input_copy.read_bytes() == original_input_bytes

    assert results["molar_basis_note"] == (
        "Molar quantities use the same basis as the input molar_flow. "
        "Use kmol as input basis to interpret mass outputs as kg and volume outputs as Nm3."
    )
    assert results["carbon_conversion_note"] == (
        "carbon_conversion represents the fraction of degradable organic carbon "
        "allocated to gas generation. The remaining degradable organic carbon is "
        "treated as associated with biological cell growth."
    )
    assert results["calculation_status"] == "Calculated successfully"
    assert results["number_of_components_read"] > 0
    assert results["CH4_Nm3"] > 0
    assert results["total_biogas_Nm3"] > 0

    workbook = load_workbook(output_file, data_only=False)

    assert "02_OUTPUTS" in workbook.sheetnames
    assert "04_EXPERIMENTAL_VALIDATION" in workbook.sheetnames
    assert "06_VALIDATION_BENCHMARKS" in workbook.sheetnames

def test_excel_reader_rejects_empty_molar_flow_in_active_row(tmp_path):
    input_copy = tmp_path / TEMPLATE_FILE.name
    shutil.copy2(TEMPLATE_FILE, input_copy)

    workbook = load_workbook(input_copy)
    worksheet = workbook["01_INPUTS"]

    header_row = None
    headers = {}

    for row in range(1, worksheet.max_row + 1):
        values = [
            worksheet.cell(row=row, column=column).value
            for column in range(1, worksheet.max_column + 1)
        ]
        normalized = [str(value).strip() if value is not None else "" for value in values]

        if "component_name" in normalized and "molar_flow" in normalized:
            header_row = row
            headers = {
                name: index + 1
                for index, name in enumerate(normalized)
                if name
            }
            break

    assert header_row is not None

    target_row = header_row + 1
    worksheet.cell(row=target_row, column=headers["active_row"]).value = True
    worksheet.cell(row=target_row, column=headers["molar_flow"]).value = None

    workbook.save(input_copy)

    with pytest.raises(ValueError, match="Molar flow is required"):
        read_biopotgas_excel(input_copy)

def test_calculate_from_excel_warns_when_global_parameter_uses_default(tmp_path):
    input_copy = tmp_path / TEMPLATE_FILE.name
    shutil.copy2(TEMPLATE_FILE, input_copy)

    workbook = load_workbook(input_copy)
    worksheet = workbook["01_INPUTS"]

    parameter_name = "carbon_conversion"
    parameter_row = None

    for row in range(4, 18):
        value = worksheet[f"A{row}"].value
        if value is not None and str(value).strip() == parameter_name:
            parameter_row = row
            break

    assert parameter_row is not None

    worksheet[f"A{parameter_row}"].value = None
    worksheet[f"B{parameter_row}"].value = None
    workbook.save(input_copy)

    results = calculate_from_excel(input_copy)

    assert results["carbon_conversion"] == 0.95
    assert (
        "Global parameter 'carbon_conversion' not found; using default 0.95."
        in results["warnings"]
    )

def test_excel_reader_requires_element_columns_for_elements_mode(tmp_path):
    input_copy = tmp_path / TEMPLATE_FILE.name
    shutil.copy2(TEMPLATE_FILE, input_copy)

    workbook = load_workbook(input_copy)
    worksheet = workbook["01_INPUTS"]

    header_row = None
    headers = {}

    for row in range(1, worksheet.max_row + 1):
        values = [
            worksheet.cell(row=row, column=column).value
            for column in range(1, worksheet.max_column + 1)
        ]
        normalized = [str(value).strip() if value is not None else "" for value in values]

        if "component_name" in normalized and "molar_flow" in normalized:
            header_row = row
            headers = {
                name: index + 1
                for index, name in enumerate(normalized)
                if name
            }
            break

    assert header_row is not None

    target_row = header_row + 1
    worksheet.cell(row=target_row, column=headers["active_row"]).value = True
    worksheet.cell(row=target_row, column=headers["input_mode"]).value = "ELEMENTS"

    worksheet.delete_cols(headers["C_atoms"])

    workbook.save(input_copy)

    with pytest.raises(ValueError, match="Missing required columns for ELEMENTS"):
        read_biopotgas_excel(input_copy)

def test_excel_reader_suggests_similar_database_component_for_typo(tmp_path):
    input_copy = tmp_path / TEMPLATE_FILE.name
    shutil.copy2(TEMPLATE_FILE, input_copy)

    workbook = load_workbook(input_copy)
    worksheet = workbook["01_INPUTS"]

    header_row = None
    headers = {}

    for row in range(1, worksheet.max_row + 1):
        values = [
            worksheet.cell(row=row, column=column).value
            for column in range(1, worksheet.max_column + 1)
        ]
        normalized = [str(value).strip() if value is not None else "" for value in values]

        if "component_name" in normalized and "molar_flow" in normalized:
            header_row = row
            headers = {
                name: index + 1
                for index, name in enumerate(normalized)
                if name
            }
            break

    assert header_row is not None

    target_row = header_row + 1
    worksheet.cell(row=target_row, column=headers["active_row"]).value = True
    worksheet.cell(row=target_row, column=headers["component_name"]).value = "GLUCOS"
    worksheet.cell(row=target_row, column=headers["input_mode"]).value = "DATABASE"

    workbook.save(input_copy)

    with pytest.raises(KeyError, match="Did you mean"):
        read_biopotgas_excel(input_copy)

def test_excel_validation_cleanup_uses_existing_sheet_extent(tmp_path):
    input_copy = tmp_path / TEMPLATE_FILE.name
    output_file = tmp_path / "calculated_output.xlsx"
    shutil.copy2(TEMPLATE_FILE, input_copy)

    workbook = load_workbook(input_copy)
    worksheet = workbook["04_EXPERIMENTAL_VALIDATION"]

    calc_header_row = None
    for row in range(1, worksheet.max_row + 1):
        values = [
            worksheet.cell(row=row, column=column).value
            for column in range(1, worksheet.max_column + 1)
        ]
        normalized = [str(value).strip() if value is not None else "" for value in values]
        if "absolute_error_CH4_Nm3" in normalized and "validation_status" in normalized:
            calc_header_row = row
            break

    assert calc_header_row is not None

    stale_row = calc_header_row + 150
    worksheet.cell(row=stale_row, column=1).value = "STALE_SAMPLE"
    worksheet.cell(row=stale_row, column=11).value = "STALE_VALUE"
    workbook.save(input_copy)

    results = calculate_from_excel(input_copy)
    write_outputs_to_excel(input_copy, output_path=output_file, results=results)

    output_workbook = load_workbook(output_file)
    output_validation = output_workbook["04_EXPERIMENTAL_VALIDATION"]

    assert output_validation.cell(row=stale_row, column=1).value is None
    assert output_validation.cell(row=stale_row, column=11).value is None

def test_excel_writer_warns_about_output_template_key_mismatches(tmp_path):
    input_copy = tmp_path / TEMPLATE_FILE.name
    output_file = tmp_path / "calculated_output.xlsx"
    shutil.copy2(TEMPLATE_FILE, input_copy)

    workbook = load_workbook(input_copy)
    worksheet = workbook["02_OUTPUTS"]

    extra_row = worksheet.max_row + 1
    worksheet.cell(row=extra_row, column=1).value = "unknown_output_key"
    workbook.save(input_copy)

    results = calculate_from_excel(input_copy)
    write_outputs_to_excel(input_copy, output_path=output_file, results=results)

    output_workbook = load_workbook(output_file)
    output_sheet = output_workbook["02_OUTPUTS"]

    warnings_value = None
    for row in range(1, output_sheet.max_row + 1):
        if output_sheet.cell(row=row, column=1).value == "warnings":
            warnings_value = output_sheet.cell(row=row, column=2).value
            break

    assert warnings_value is not None
    assert "Calculated result keys not written to 02_OUTPUTS" in warnings_value
    assert "02_OUTPUTS keys not found in calculated results" in warnings_value
    assert "unknown_output_key" in warnings_value